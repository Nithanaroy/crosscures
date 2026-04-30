"""Load Kapil's eval Excel into DuckDB eval tables."""
import duckdb
import openpyxl
from offline_eval.config import DUCKDB_PATH, EXCEL_PATH


def _parse_note_ids(raw: str) -> list:
    """Parse semicolon-delimited note IDs into a list of ints."""
    if not raw or not str(raw).strip():
        return []
    return [int(x.strip()) for x in str(raw).split(";") if x.strip()]


def load_eval_data(db_path: str = None, excel_path: str = None):
    """Read the 3 Excel sheets and insert into DuckDB as eval tables."""
    db_path = db_path or str(DUCKDB_PATH)
    excel_path = excel_path or str(EXCEL_PATH)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    con = duckdb.connect(db_path)

    # --- eval_cases ---
    con.execute("DROP TABLE IF EXISTS eval_cases")
    con.execute("""
        CREATE TABLE eval_cases (
            case_id           TEXT PRIMARY KEY,
            patient_id        BIGINT,
            cutoff_date       DATE,
            case_domain       TEXT,
            case_name         TEXT,
            note_ids_included BIGINT[],
            input_summary     TEXT
        )
    """)

    ws = wb["input reference"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        case_id, patient_id, cutoff_date, note_ids_raw, case_domain, case_name, input_summary = row[:7]
        if not case_id:
            continue
        note_ids = _parse_note_ids(note_ids_raw)
        cutoff_str = cutoff_date.strftime("%Y-%m-%d") if hasattr(cutoff_date, "strftime") else str(cutoff_date)[:10]
        con.execute(
            "INSERT INTO eval_cases VALUES (?, ?, ?, ?, ?, ?, ?)",
            [str(case_id), int(patient_id), cutoff_str, str(case_domain), str(case_name), note_ids, str(input_summary or "")]
        )

    # --- history_targets ---
    con.execute("DROP TABLE IF EXISTS history_targets")
    con.execute("""
        CREATE TABLE history_targets (
            target_id          TEXT PRIMARY KEY,
            case_id            TEXT,
            patient_id         BIGINT,
            case_domain        TEXT,
            concept            TEXT,
            target_slot        TEXT,
            patient_answerable TEXT,
            severity           TEXT,
            weight             INTEGER,
            source_note_ids    BIGINT[],
            rationale          TEXT
        )
    """)

    ws = wb["question list "]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    target_idx = 0
    for row in rows:
        case_id = row[0]
        if not case_id or not str(case_id).strip():
            continue
        patient_id, case_domain, concept, target_slot = row[1], row[2], row[3], row[4]
        patient_answerable, severity, weight, source_note_ids_raw, rationale = row[5], row[6], row[7], row[8], row[9]
        note_ids = _parse_note_ids(source_note_ids_raw)
        target_id = f"{case_id}_T{target_idx:03d}"
        target_idx += 1
        con.execute(
            "INSERT INTO history_targets VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [target_id, str(case_id), int(patient_id), str(case_domain), str(concept),
             str(target_slot), str(patient_answerable or "yes"), str(severity),
             int(weight or 1), note_ids, str(rationale or "")]
        )

    # --- linked_outcomes ---
    con.execute("DROP TABLE IF EXISTS linked_outcomes")
    con.execute("""
        CREATE TABLE linked_outcomes (
            outcome_id      TEXT PRIMARY KEY,
            case_id         TEXT,
            patient_id      BIGINT,
            case_domain     TEXT,
            concept         TEXT,
            event_type      TEXT,
            severity        TEXT,
            source_note_ids BIGINT[],
            outcome_detail  TEXT
        )
    """)

    ws = wb["outcomes references"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    outcome_idx = 0
    for row in rows:
        case_id = row[0]
        if not case_id or not str(case_id).strip():
            continue
        patient_id, case_domain, concept, event_type, severity, source_note_ids_raw, outcome_detail = row[1:8]
        note_ids = _parse_note_ids(source_note_ids_raw)
        outcome_id = f"{case_id}_O{outcome_idx:03d}"
        outcome_idx += 1
        con.execute(
            "INSERT INTO linked_outcomes VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [outcome_id, str(case_id), int(patient_id), str(case_domain), str(concept),
             str(event_type), str(severity), note_ids, str(outcome_detail or "")]
        )

    con.close()
    wb.close()

    print(f"[DONE] Loaded eval data into {db_path}")
    _print_summary(db_path)


def _print_summary(db_path: str):
    con = duckdb.connect(db_path, read_only=True)
    cases = con.execute("SELECT COUNT(*) FROM eval_cases").fetchone()[0]
    targets = con.execute("SELECT COUNT(*) FROM history_targets").fetchone()[0]
    outcomes = con.execute("SELECT COUNT(*) FROM linked_outcomes").fetchone()[0]
    print(f"  eval_cases:      {cases}")
    print(f"  history_targets: {targets}")
    print(f"  linked_outcomes: {outcomes}")

    print("\n  Cases by domain:")
    for row in con.execute("SELECT case_domain, COUNT(*) FROM eval_cases GROUP BY case_domain ORDER BY COUNT(*) DESC").fetchall():
        print(f"    {row[0]:25} {row[1]}")
    con.close()


if __name__ == "__main__":
    load_eval_data()
